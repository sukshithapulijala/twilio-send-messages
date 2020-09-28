import openpyxl, twilio.rest, os
from openpyxl.utils.exceptions import SheetTitleException, InvalidFileException


class MessageForwarding:

    def __init__(self, account_sid, auth_token, my_cell_phone, my_twilio_number):
        self.account_sid = account_sid
        self.auth_token = auth_token
        self.twilio_cli = twilio.rest.Client(self.account_sid, self.auth_token)
        self.my_twilio_number = my_twilio_number
        self.my_cell_phone = my_cell_phone
        self.blank_count = 0
        self.students = []

    def message(self):

        send_str = str("The following message will be sent to %s:\n" % self.my_cell_phone)
        send_str = send_str + str("Our records show that you have %d missing grades, for the following students:\n" % self.blank_count)
        
        # append students name to string for output
        for x in self.students:
            send_str = send_str + str(x) + str("\n")
        return send_str

    def worksheet(self, excel_file):

        # handling invalid file exceptions
        try:
            wb = openpyxl.load_workbook(excel_file)
            wb.get_sheet_names()
            sheet = wb.get_sheet_by_name(wb.sheetnames[0])
            rows = sheet.max_row
            columns = sheet.max_column
            names = []

            # iterate through the sheet leaving first row and column and count blank cells
            for iter_row in range(1, rows + 1):
                for iter_column in range(1, columns + 1):
                    if sheet.cell(row=iter_row, column=iter_column).value is None:
                        self.blank_count += 1
                        names.append(sheet['A' + str(iter_row)].value)

            # removing duplicates from list names
            self.students = list(dict.fromkeys(names))
            print("You have %d missing grades, for the following students:" % self.blank_count)

            for name in self.students:
                print(name)

            # send the output as a text message
            mf.send_msg()

        except SheetTitleException:
            print("Invalid excel file. Check your file name and try again.")

        except InvalidFileException:
            print("Invalid file format. Check your file name and try again.")

    # Send string created in message() to a mobile number
    def send_msg(self):
        self.twilio_cli.messages.create(body=self.message()
                                        , from_=self.my_twilio_number, to=self.my_cell_phone)


if __name__ == "__main__":
    if len(os.sys.argv) == 6:
        test_excel_file = os.sys.argv[1]
        mf = MessageForwarding(os.sys.argv[2], os.sys.argv[3], os.sys.argv[4], os.sys.argv[5])
        mf.worksheet(test_excel_file)
    else:
        print("Run as: p3.py <test_excel_file> <account_sid> <auth_token> <cell_phone_number> <twilio_number>")
